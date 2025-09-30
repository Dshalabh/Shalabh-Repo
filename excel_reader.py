from openpyxl import load_workbook
import sys

def read_excel_file(file_path):
    try:
        # Load the Excel file
        workbook = load_workbook(file_path, data_only=True)

        print("Excel file contents:")
        print("=" * 50)

        for sheet_name in workbook.sheetnames:
            print(f"\nSheet: {sheet_name}")
            print("-" * 30)

            worksheet = workbook[sheet_name]

            # Get all data from the sheet
            for row in worksheet.iter_rows():
                row_data = []
                for cell in row:
                    value = cell.value if cell.value is not None else ""
                    row_data.append(str(value))
                if any(row_data):  # Only print non-empty rows
                    print("\t".join(row_data))

            print("\n")

    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    file_path = r"C:\Users\ShalabDo\Documents\Competative Analysis.xlsx"
    read_excel_file(file_path)